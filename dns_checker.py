#!/usr/bin/env python3
# Copyright 2026 Sigurd Felix
# SPDX-License-Identifier: Apache-2.0
"""
DNS Checker
===========
Controleert DNS-configuratie van domeinen op e-mailbeveiliging en infrastructuur.
Leest domeinen uit een domains.txt (één domein per regel) of via het opstartmenu.
Output: Excel-werkmap of HTML-rapport met timestamp.

Vereisten:
    pip install dnspython requests openpyxl

Gebruik:
    python dns_checker.py                          # opstartmenu
    python dns_checker.py --domain mijnbedrijf.nl  # enkel domein
    python dns_checker.py --domains domains.txt    # tekstbestand met domeinen
    python dns_checker.py --workers 10             # parallelle threads
    python dns_checker.py --timeout 5              # DNS timeout in seconden
    python dns_checker.py --http-timeout 5         # HTTP timeout in seconden
    python dns_checker.py --insecure               # TLS-waarschuwingen onderdrukken
    python dns_checker.py --no-input               # geen interactieve vragen
"""

# ─── Dependency check (moet als eerste, vóór andere imports) ─────────────────
import sys
import subprocess

VEREISTE_PAKKETTEN = {
    "dns":      "dnspython",
    "requests": "requests",
    "openpyxl": "openpyxl",
}

def _controleer_en_installeer_pakketten():
    ontbrekend = []
    for import_naam, pip_naam in VEREISTE_PAKKETTEN.items():
        try:
            __import__(import_naam)
        except ImportError:
            ontbrekend.append((import_naam, pip_naam))

    if not ontbrekend:
        return  # alles aanwezig

    print()
    print("╔══════════════════════════════════════════════════╗")
    print("║           Ontbrekende pakketten gevonden         ║")
    print("╚══════════════════════════════════════════════════╝")
    print()
    print("  De volgende Python-pakketten zijn vereist maar")
    print("  nog niet geïnstalleerd:")
    print()
    for _, pip_naam in ontbrekend:
        print(f"    • {pip_naam}")
    print()

    antwoord = input("  Nu automatisch installeren? (j/n): ").strip().lower()
    if antwoord not in ("j", "ja", "y", "yes"):
        print()
        print("  Installeer de pakketten handmatig met:")
        pip_namen = " ".join(p for _, p in ontbrekend)
        print(f"    pip install {pip_namen}")
        print()
        input("  Druk Enter om af te sluiten...")
        sys.exit(1)

    print()
    for import_naam, pip_naam in ontbrekend:
        print(f"  📦 Installeer {pip_naam}...")
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pip_naam],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            print(f"  ✅ {pip_naam} geïnstalleerd.")
        except subprocess.CalledProcessError:
            print(f"  ❌ Installatie van {pip_naam} mislukt.")
            print(f"     Probeer handmatig: pip install {pip_naam}")
            print()
            input("  Druk Enter om af te sluiten...")
            sys.exit(1)

    print()
    print("  ✅ Alle pakketten geïnstalleerd. Script wordt gestart...")
    print()

_controleer_en_installeer_pakketten()

# ─── Overige imports (na dependency check) ───────────────────────────────────
import os

import csv
import re
import time
import argparse
import logging
import random
from datetime import datetime
from urllib.parse import urlparse, quote_plus
from concurrent.futures import ThreadPoolExecutor, as_completed

import dns.resolver
import dns.dnssec
import dns.name
import dns.query
import dns.flags
import requests



# ─── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─── Constanten ─────────────────────────────────────────────────────────────
DNS_TIMEOUT  = 3    # seconden per DNS-query
HTTP_TIMEOUT = 5    # seconden voor HTTP-requests (security.txt, MTA-STS)
INSECURE     = False  # TLS-waarschuwingen onderdrukken (--insecure flag)
MAX_WORKERS  = 10   # parallelle threads

MICROSOFT_MX_PATTERNS = [
    "mail.protection.outlook.com",   # klassiek EOP
    "mx.microsoft",                  # nieuw Microsoft MX (2024+)
    "outlook.com",
]
GOOGLE_MX_PATTERNS = [
    "aspmx.l.google.com",
    "googlemail.com",
    "smtp.google.com",
]

# Bekende Secure Mail Gateways (SMG) — zitten vóór de echte mailprovider
SMG_MX_PATTERNS = {
    "mimecast.com":        "Mimecast",
    "mimecast9.com":       "Mimecast",
    "ppe-hosted.com":      "Proofpoint",
    "pphosted.com":        "Proofpoint",
    "proofpoint.com":      "Proofpoint",
    "barracudanetworks.com": "Barracuda",
    "cudamail.com":        "Barracuda",
    "hornetsecurity.com":  "Hornetsecurity",
    "spamexperts.com":     "SpamExperts",
    "spamexperts.eu":      "SpamExperts",
    "antispameurope.com":  "AntiSpamEurope",
    "mailroute.net":       "MailRoute",
    "messagelabs.com":     "Symantec/MessageLabs",
    "emailsrvr.com":       "Rackspace",
    "sophos.com":          "Sophos",
    "ironport.com":        "Cisco IronPort",
    "trendmicro.com":      "Trend Micro",
    "forcepoint.com":      "Forcepoint",
}

MICROSOFT_SPF_INCLUDES = [
    "spf.protection.outlook.com",
    "protection.outlook.com",
    "microsoft.com",
    "sharepointonline.com",
    "spf.messaging.microsoft.com",   # nieuw Microsoft SPF domein
]
GOOGLE_SPF_INCLUDES = [
    "_spf.google.com",
    "google.com",
    "_netblocks.google.com",
]

# SMG-specifieke SPF includes — als deze aanwezig zijn, scan dan dieper
SMG_SPF_PATTERNS = {
    "mimecast.com":        "Mimecast",
    "ppe-hosted.com":      "Proofpoint",
    "pphosted.com":        "Proofpoint",
    "proofpoint.com":      "Proofpoint",
    "barracudanetworks.com": "Barracuda",
    "cudamail.com":        "Barracuda",
    "hornetsecurity.com":  "Hornetsecurity",
    "spamexperts.com":     "SpamExperts",
    "spamexperts.eu":      "SpamExperts",
    "antispameurope.com":  "AntiSpamEurope",
}

DKIM_MICROSOFT_SELECTORS = ["selector1", "selector2"]
DKIM_GOOGLE_SELECTORS    = ["google", "google2"]

KNOWN_DNS_PROVIDERS = {
    "cloudflare": "Cloudflare",
    "azure-dns":  "Azure DNS",
    "transip":    "TransIP",
    "sidn":       "SIDN",
    "mijndomein": "MijnDomein",
    "yourhosting": "YourHosting",
    "versio":     "Versio",
    "antagonist": "Antagonist",
    "byte":       "Byte",
    "combell":    "Combell",
    "awsdns":     "AWS Route53",
    "googledomains": "Google Domains",
    "googledns":  "Google DNS",
    "ns1.com":    "NS1",
    "dnsimple":   "DNSimple",
}


# ─── Hulpfuncties ───────────────────────────────────────────────────────────

def make_resolver(timeout: int = DNS_TIMEOUT) -> dns.resolver.Resolver:
    r = dns.resolver.Resolver()
    r.lifetime = timeout
    r.timeout  = timeout
    return r


def extract_domain(url: str) -> str | None:
    """Haal het kale domein (zonder protocol, www, pad) uit een URL."""
    if not url:
        return None
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    try:
        parsed = urlparse(url)
        host = parsed.hostname or ""
        # Verwijder www. prefix
        if host.startswith("www."):
            host = host[4:]
        return host.lower() if host else None
    except Exception:
        return None


def resolve(resolver, domain: str, rtype: str) -> list[str]:
    """Geeft een lijst van strings terug, of [] bij mislukking."""
    try:
        answers = resolver.resolve(domain, rtype)
        return [str(r).rstrip(".") for r in answers]
    except dns.resolver.NXDOMAIN:
        return []      # domein bestaat niet
    except dns.resolver.NoAnswer:
        return []      # domein bestaat wel, maar geen record van dit type
    except dns.resolver.Timeout:
        return ["TIMEOUT"]
    except Exception:
        return []


def txt_records(resolver, domain: str) -> list[str]:
    """
    Haalt TXT-records op. Een DNS TXT-record mag bestaan uit meerdere
    deelstrings (elk max 255 bytes) die logisch één geheel vormen.
    Die worden hier per rdata samengevoegd met een spatie als scheidingsteken,
    zodat langere SPF-records intact blijven.

    Valt terug op rdata.to_text() bij exotische encodings en logt decode-issues.
    """
    try:
        answers = resolver.resolve(domain, "TXT")
        result = []
        for rdata in answers:
            delen = []
            for b in rdata.strings:
                if isinstance(b, bytes):
                    try:
                        delen.append(b.decode("utf-8"))
                    except UnicodeDecodeError:
                        try:
                            delen.append(b.decode("latin-1"))
                            log.debug(f"TXT decode latin-1 fallback voor {domain}")
                        except Exception:
                            # Laatste redmiddel: via rdata string representatie
                            tekst = rdata.to_text().strip('"')
                            log.debug(f"TXT decode via to_text() fallback voor {domain}")
                            delen.append(tekst)
                            break  # to_text() geeft het hele record, niet loopen
                else:
                    delen.append(str(b).strip('"'))
            result.append(" ".join(delen))
        return result
    except Exception:
        return []


# ─── Checks ─────────────────────────────────────────────────────────────────

def check_a_aaaa(resolver, domain):
    a    = resolve(resolver, domain, "A")
    aaaa = resolve(resolver, domain, "AAAA")
    return {
        "a_records":   "; ".join(a)    if a    else "",
        "aaaa_records": "; ".join(aaaa) if aaaa else "",
        "ipv6_support": "yes" if aaaa else "no",
    }


def check_mx(resolver, domain):
    """MX records ophalen met detectie van Microsoft, Google en SMG's.
    Probeert maximaal 3 keer voor het opgeeft (DNS is soms instabiel)."""
    hosts = []
    for poging in range(1, 4):
        try:
            answers = resolver.resolve(domain, "MX")
            mx_entries = sorted(answers, key=lambda r: r.preference)
            hosts = [
                f"{r.preference} {str(r.exchange).rstrip('.').lower()}"
                for r in mx_entries
            ]
            if hosts:
                break
        except dns.resolver.NoAnswer:
            break   # domein bestaat maar heeft echt geen MX, niet opnieuw proberen
        except dns.resolver.NXDOMAIN:
            break   # domein bestaat niet
        except Exception:
            if poging < 3:
                time.sleep(1)
            continue

    if not hosts:
        # Fallback: kijk of mail.domein een A-record heeft
        try:
            fallback = resolver.resolve(f"mail.{domain}", "A")
            if fallback:
                return {
                    "mx_records":    f"mail.{domain} (via A-record fallback)",
                    "mail_provider": "onbekend",
                    "smg":           "",
                }
        except Exception:
            pass
        return {"mx_records": "", "mail_provider": "", "smg": ""}

    # Detecteer SMG
    smg_naam = ""
    for h in hosts:
        for pattern, naam in SMG_MX_PATTERNS.items():
            if pattern in h:
                smg_naam = naam
                break
        if smg_naam:
            break

    # Detecteer eindbestemming via MX — definitieve provider wordt na SPF bepaald
    mx_provider = "onbekend"
    for h in hosts:
        if any(p in h for p in MICROSOFT_MX_PATTERNS):
            mx_provider = "Microsoft 365"
            break
        if any(p in h for p in GOOGLE_MX_PATTERNS):
            mx_provider = "Google Workspace"
            break

    if smg_naam and mx_provider == "onbekend":
        mx_provider = f"via {smg_naam} (eindbestemming onbekend)"

    return {
        "mx_records":    "; ".join(hosts),
        "mail_provider": mx_provider,   # wordt later overschreven na SPF
        "smg":           smg_naam,
    }


def _spf_eindbestemming(spf: str) -> str:
    """
    Bepaal de mail-eindbestemming puur op basis van SPF includes.
    Kijkt naar Microsoft- en Google-specifieke includes.
    """
    if any(inc in spf for inc in MICROSOFT_SPF_INCLUDES):
        return "Microsoft 365"
    if any(inc in spf for inc in GOOGLE_SPF_INCLUDES):
        return "Google Workspace"
    return ""


def _spf_uitklappen(resolver, domain: str, bezocht: set = None, diepte: int = 0) -> dict:
    """
    Volgt SPF-includes recursief en verzamelt:
    - totaal aantal DNS-lookups (RFC max 10)
    - alle unieke includes/mechanismen
    - effectieve verzenders (leesbare providers)
    Geeft dict terug met lookups, verzenders, foutmeldingen.
    """
    if bezocht is None:
        bezocht = set()
    if domain in bezocht or diepte > 10:
        return {"lookups": 0, "verzenders": [], "fouten": []}
    bezocht.add(domain)

    try:
        records = txt_records(resolver, domain)
    except Exception as e:
        return {"lookups": 0, "verzenders": [], "fouten": [f"DNS fout bij {domain}: {e}"]}

    spf_records = [r for r in records if r.startswith("v=spf1")]
    if not spf_records:
        return {"lookups": 0, "verzenders": [], "fouten": []}

    spf = spf_records[0]

    # Bekende provider-labels op basis van include-domein
    PROVIDER_LABELS = {
        # Microsoft
        "protection.outlook.com":   "Microsoft 365",
        "spf.protection.outlook":   "Microsoft 365",
        "mx.microsoft":             "Microsoft 365",
        # Google
        "google.com":               "Google Workspace",
        "googlemail.com":           "Google Workspace",
        # SMG / filters
        "mimecast.com":             "Mimecast (SMG)",
        "ppe-hosted.com":           "Proofpoint (SMG)",
        "pphosted.com":             "Proofpoint (SMG)",
        "barracudanetworks.com":    "Barracuda (SMG)",
        "hornetsecurity.com":       "Hornetsecurity (SMG)",
        "spamexperts.com":          "SpamExperts (SMG)",
        "mailprotect.be":           "Mailprotect (SMG)",
        "spamfilter.nl":            "SpamFilter.nl (SMG)",
        # Transactionele/marketing mail
        "sendgrid.net":             "SendGrid (Twilio)",
        "mailgun.org":              "Mailgun",
        "amazonses.com":            "Amazon SES",
        "mailchimp.com":            "Mailchimp",
        "mandrillapp.com":          "Mandrill (Mailchimp)",
        "exacttarget.com":          "Salesforce Marketing Cloud",
        "salesforce.com":           "Salesforce",
        "mktomail.com":             "Marketo",
        "hubspot.com":              "HubSpot",
        "zendesk.com":              "Zendesk",
        "flowmailer.com":           "Flowmailer",
        "flowmailer.net":           "Flowmailer",
        "mailjet.com":              "Mailjet",
        "sendinblue.com":           "Brevo (Sendinblue)",
        "brevo.com":                "Brevo",
        "postmarkapp.com":          "Postmark",
        "sparkpostmail.com":        "SparkPost",
        "cmail1.com":               "Campaign Monitor",
        "cmail2.com":               "Campaign Monitor",
        "createsend.com":           "Campaign Monitor",
        "constantcontact.com":      "Constant Contact",
        "klaviyo.com":              "Klaviyo",
        "intercom.io":              "Intercom",
        "freshdesk.com":            "Freshdesk",
        "helpscoutemail.com":       "Help Scout",
        "zoho.com":                 "Zoho",
        "msg91.com":                "MSG91",
        "smtp.com":                 "SMTP.com",
        # Hostingproviders NL
        "transip.nl":               "TransIP",
        "transip.net":              "TransIP",
        "mijndomein.nl":            "MijnDomein",
        "yourhosting.nl":           "YourHosting",
        "antagonist.nl":            "Antagonist",
        "byte.nl":                  "Byte",
        "neostrada.nl":             "Neostrada",
        "one.com":                  "One.com",
        "amen.fr":                  "Amen",
        "combell.com":              "Combell",
    }

    # Tel mechanismen die een DNS-lookup kosten
    mechanismen = re.findall(r'\b(include|a|mx|ptr|exists):([^\s]+)', spf)
    redirect     = re.search(r'redirect=([^\s]+)', spf)
    lookup_count = len(mechanismen) + (1 if redirect else 0)

    verzenders = []
    fouten     = []
    sub_totaal = 0

    for mech, waarde in mechanismen:
        if mech == "include":
            # Label bepalen
            label = next((v for k, v in PROVIDER_LABELS.items() if k in waarde), None)
            if label and label not in verzenders:
                verzenders.append(label)
            # Recursief uitklappen
            sub = _spf_uitklappen(resolver, waarde, bezocht, diepte + 1)
            sub_totaal += sub["lookups"]
            for v in sub["verzenders"]:
                if v not in verzenders:
                    verzenders.append(v)
            fouten.extend(sub["fouten"])

    if redirect:
        sub = _spf_uitklappen(resolver, redirect.group(1), bezocht, diepte + 1)
        sub_totaal += sub["lookups"]
        for v in sub["verzenders"]:
            if v not in verzenders:
                verzenders.append(v)
        fouten.extend(sub["fouten"])

    return {
        "lookups":   lookup_count + sub_totaal,
        "verzenders": verzenders,
        "fouten":    fouten,
    }


def check_spf(resolver, domain):
    records = txt_records(resolver, domain)
    alle_spf = [r for r in records if r.startswith("v=spf1")]

    # Dubbele SPF-records — RFC 7208 staat er maar één toe
    dubbel = len(alle_spf) > 1

    if not alle_spf:
        return {
            "spf_record":             "",
            "spf_strictness":         "",
            "spf_mail_provider_hint": "",
            "spf_smg":                "",
            "spf_lookups":            "",
            "spf_record_lengte":      "",
            "spf_verzenders":         "",
            "spf_fout":               "",
        }

    spf = alle_spf[0]

    # Recordlengte — RFC 7208 vereist dat het samengevoegde SPF-record door de
    # ontvangende kant correct verwerkt kan worden. Eén TXT-string is intern max
    # 255 bytes, maar een record mag meerdere strings bevatten die samengevoegd
    # worden. We meten de totale lengte van het samengevoegde record als indicator.
    lengte       = len(spf.encode("utf-8"))
    lengte_label = f"{lengte} bytes" + (" ⚠️ >255" if lengte > 255 else " ✓")

    # Bij redirect= is het doelrecord leidend voor strictheid en verzenders
    redirect_match = re.search(r'\bredirect=([^\s]+)', spf)
    spf_voor_strictheid = spf
    redirect_domein = None
    if redirect_match:
        redirect_domein = redirect_match.group(1)
        try:
            doel_records = txt_records(resolver, redirect_domein)
            doel_spf = next((r for r in doel_records if r.startswith("v=spf1")), None)
            if doel_spf:
                spf_voor_strictheid = doel_spf
        except Exception:
            pass

    # Strictheid — gebruik regex zodat bijv. "ip4:1.2.3.4/~all" niet vals matcht
    if re.search(r'\+all\b', spf_voor_strictheid):
        strictness = "+all ⚠️"
    elif re.search(r'\?all\b', spf_voor_strictheid):
        strictness = "?all ⚠️"
    elif re.search(r'-all\b', spf_voor_strictheid):
        strictness = "-all (hard fail)"
    elif re.search(r'~all\b', spf_voor_strictheid):
        strictness = "~all (soft fail)"
    elif redirect_domein and spf_voor_strictheid == spf:
        # redirect aanwezig maar doelrecord niet opgehaald
        strictness = f"redirect → {redirect_domein} (niet opgehaald)"
    elif redirect_domein:
        strictness = f"redirect → {redirect_domein} (geen all-modifier)"
    else:
        strictness = "geen all-modifier ⚠️"

    # Detecteer SMG in SPF
    spf_smg = ""
    for pattern, naam in SMG_SPF_PATTERNS.items():
        if pattern in spf:
            spf_smg = naam
            break

    # Recursief uitklappen — lookups, verzenders, fouten
    uitgeklapt   = _spf_uitklappen(resolver, domain)
    totaal_lookups = uitgeklapt["lookups"]
    verzenders   = uitgeklapt["verzenders"]
    spf_fouten   = uitgeklapt["fouten"]

    lookup_label = str(totaal_lookups) + (" ⚠️ >10" if totaal_lookups > 10 else " ✓")

    # Fouten samenvoegen
    fout_parts = []
    if dubbel:
        fout_parts.append(f"⚠️ {len(alle_spf)} SPF-records gevonden — RFC staat er maar 1 toe")
    if spf_fouten:
        fout_parts.extend(spf_fouten)

    # Eindbestemming hint (voor mail_provider kolom)
    hint = _spf_eindbestemming(spf)
    if spf_smg and not hint:
        includes = re.findall(r"include:([^\s]+)", spf)
        for inc_domain in includes:
            if any(p in inc_domain for p in SMG_SPF_PATTERNS):
                continue
            try:
                sub_records = txt_records(resolver, inc_domain)
                sub_spf = next((r for r in sub_records if r.startswith("v=spf1")), None)
                if sub_spf:
                    hint = _spf_eindbestemming(sub_spf)
                    if hint:
                        hint += f" (via {inc_domain})"
                        break
            except Exception:
                continue
        if not hint:
            hint = f"onbekend (via {spf_smg})"
    if not hint:
        hint = "other"

    return {
        "spf_record":             spf,
        "spf_strictness":         strictness,
        "spf_mail_provider_hint": hint,
        "spf_smg":                spf_smg,
        "spf_lookups":            lookup_label,
        "spf_record_lengte":      lengte_label,
        "spf_verzenders":         "; ".join(verzenders) if verzenders else "onbekend",
        "spf_fout":               "; ".join(fout_parts),
    }


def check_dmarc(resolver, domain):
    records = txt_records(resolver, f"_dmarc.{domain}")
    dmarc = next((r for r in records if r.startswith("v=DMARC1")), None)
    if not dmarc:
        return {
            "dmarc_record":  "",
            "dmarc_policy":  "none",
            "dmarc_rua":     "",
            "dmarc_pct":     "",
        }

    m_policy = re.search(r"\bp=(\w+)", dmarc)
    policy   = m_policy.group(1).lower() if m_policy else "unknown"

    m_rua = re.search(r"\brua=([^\s;]+)", dmarc)
    m_pct = re.search(r"\bpct=(\d+)", dmarc)

    rua = m_rua.group(1) if m_rua else ""
    pct = m_pct.group(1) if m_pct else "100"

    return {
        "dmarc_record":  dmarc,
        "dmarc_policy":  policy,
        "dmarc_rua":     rua,
        "dmarc_pct":     pct,
    }


def check_dkim(resolver, domain):
    results = {}

    def _check_selector(sel: str) -> str:
        """Geeft found/not found terug voor een selector.
        Volgt CNAME als er geen directe TXT is (Microsoft 365 gebruikt dit)."""
        host = f"{sel}._domainkey.{domain}"
        txts = txt_records(resolver, host)
        if not txts:
            try:
                cname_answers = resolver.resolve(host, "CNAME")
                cname_target  = str(cname_answers[0].target).rstrip(".")
                txts = txt_records(resolver, cname_target)
            except Exception:
                pass
        if not txts:
            return "not found"
        record = " ".join(txts)
        m = re.search(r"p=([A-Za-z0-9+/=]+)", record)
        if not m or not m.group(1):
            return "found (geen sleutel)"
        return "found"

    for sel in DKIM_MICROSOFT_SELECTORS:
        results[f"dkim_{sel}"] = _check_selector(sel)

    for sel in DKIM_GOOGLE_SELECTORS:
        results[f"dkim_{sel}"] = _check_selector(sel)

    ms_found = any(results[f"dkim_{s}"] == "found" for s in DKIM_MICROSOFT_SELECTORS)
    g_found  = any(results[f"dkim_{s}"] == "found" for s in DKIM_GOOGLE_SELECTORS)

    results["dkim_microsoft"] = "found" if ms_found else "not found"
    results["dkim_google"]    = "found" if g_found  else "not found"
    return results


def check_ns(resolver, domain):
    ns_raw = resolve(resolver, domain, "NS")
    ns_list = [r.lower() for r in ns_raw]

    provider = "unknown"
    for ns in ns_list:
        for pattern, name in KNOWN_DNS_PROVIDERS.items():
            if pattern in ns:
                provider = name
                break
        if provider != "unknown":
            break

    return {
        "ns_records":    "; ".join(ns_list),
        "dns_provider":  provider,
    }


def check_dnssec(resolver, domain):
    """Simpele DNSSEC check: kijkt of er DNSKEY records aanwezig zijn."""
    try:
        resolver.resolve(domain, "DNSKEY")
        return {"dnssec": "yes"}
    except Exception:
        return {"dnssec": "no"}


def check_mta_sts(resolver, domain):
    """
    Controleert MTA-STS in twee stappen:
    1. DNS TXT-record _mta-sts.{domain} aanwezig?
    2. Policy-bestand ophalen en 'mode' bepalen (enforce / testing / none)
    """
    records = txt_records(resolver, f"_mta-sts.{domain}")
    if not any("v=STSv1" in r for r in records):
        return {"mta_sts": "not found", "mta_sts_mode": ""}

    # Haal het policy-bestand op — verify=True want MTA-STS beoogt juist TLS-handhaving.
    # Een ongeldig certificaat is hier een bevindingswaardig probleem, geen stille fout.
    url = f"https://mta-sts.{domain}/.well-known/mta-sts.txt"
    try:
        resp = requests.get(url, timeout=HTTP_TIMEOUT, verify=True,
                            allow_redirects=True,
                            headers={"User-Agent": "DNS-Checker/1.0"})
        if resp.status_code == 200:
            m = re.search(r"mode\s*:\s*(\w+)", resp.text, re.IGNORECASE)
            mode = m.group(1).lower() if m else "onbekend"
        else:
            mode = f"policy niet bereikbaar (HTTP {resp.status_code})"
    except requests.exceptions.SSLError:
        mode = "TLS-certificaat ongeldig ⚠️"
    except requests.exceptions.ConnectionError:
        mode = "policy niet bereikbaar (verbindingsfout)"
    except requests.exceptions.Timeout:
        mode = "policy niet bereikbaar (timeout)"
    except Exception:
        mode = "policy niet bereikbaar"

    return {"mta_sts": "found", "mta_sts_mode": mode}


def check_bimi(resolver, domain):
    records = txt_records(resolver, f"default._bimi.{domain}")
    found = any("v=BIMI1" in r for r in records)
    return {"bimi": "found" if found else "not found"}


def check_security_txt(domain):
    """Probeert security.txt op te halen via HTTPS en HTTP."""
    urls = [
        f"https://{domain}/.well-known/security.txt",
        f"https://{domain}/security.txt",
        f"http://{domain}/.well-known/security.txt",
    ]
    for url in urls:
        try:
            resp = requests.get(url, timeout=HTTP_TIMEOUT, verify=not INSECURE,
                                allow_redirects=True,
                                headers={"User-Agent": "DNS-Checker/1.0"})
            if resp.status_code == 200 and "Contact:" in resp.text:
                return {
                    "security_txt":     "found",
                    "security_txt_url": url,
                }
        except Exception:
            continue
    return {"security_txt": "not found", "security_txt_url": ""}


def check_caa(resolver, domain):
    """
    Controleert CAA-records (Certificate Authority Authorization).
    CAA bepaalt welke CA's SSL-certificaten mogen uitgeven voor het domein.
    """
    try:
        answers = resolver.resolve(domain, "CAA")
        caa_list = []
        for r in answers:
            # CAA record: <flag> <tag> <value>
            caa_list.append(str(r).strip('"'))
        return {"caa_records": "; ".join(caa_list)}
    except Exception:
        return {"caa_records": ""}


# ─── Tip generator ──────────────────────────────────────────────────────────

def genereer_tips(result: dict) -> str:
    """
    Analyseert de DNS-resultaten en genereert concrete, beknopte adviezen.
    Geeft een puntkomma-gescheiden string terug voor de CSV-kolom.
    """
    tips = []

    spf        = result.get("spf_record", "")
    spf_strict = result.get("spf_strictness", "")
    spf_fout   = result.get("spf_fout", "")
    spf_lookups = result.get("spf_lookups", "")
    dmarc      = result.get("dmarc_record", "")
    dmarc_p    = result.get("dmarc_policy", "")
    dmarc_rua  = result.get("dmarc_rua", "")
    dnssec     = result.get("dnssec", "")

    # ── SPF ──────────────────────────────────────────────────────────────────
    if not spf:
        tips.append("SPF ontbreekt: iedereen kan e-mail versturen namens dit domein. Voeg een SPF-record toe.")
    else:
        if "+all" in spf_strict:
            tips.append("SPF staat op +all: elke server mag mailen namens dit domein — dit biedt geen enkele bescherming. Verander naar -all of ~all.")
        elif "?all" in spf_strict:
            tips.append("SPF staat op ?all (neutraal): ontvangende servers doen niets met niet-geautoriseerde verzenders. Verander naar -all of ~all.")
        elif "geen all" in spf_strict and "redirect" not in spf_strict:
            tips.append("SPF heeft geen afsluitende all-modifier: het record is onvolledig. Voeg -all of ~all toe aan het einde.")
        if "⚠️ >10" in spf_lookups:
            tips.append(f"SPF overschrijdt het DNS-lookuplimiet van 10 ({spf_lookups}): ontvangende servers mogen de SPF-check laten mislukken. Verminder het aantal includes.")
        spf_lengte = result.get("spf_record_lengte", "")
        if "⚠️" in spf_lengte:
            tips.append(f"SPF-record is lang ({spf_lengte}): het samengevoegde record overschrijdt 255 bytes. Overweeg includes te consolideren of een redirect= te gebruiken (RFC 7208).")
        if spf_fout:
            for f in spf_fout.split(";"):
                f = f.strip()
                if f:
                    tips.append(f"SPF: {f}")
    # -all en ~all zijn beide prima, geen tip nodig

    # ── DMARC ────────────────────────────────────────────────────────────────
    if not dmarc:
        tips.append("DMARC ontbreekt: domeinspoofing voor phishing is onbeperkt mogelijk. Voeg minimaal p=none met rua toe om inzicht te krijgen.")
    else:
        if dmarc_p == "none":
            tips.append("DMARC p=none: het record is aanwezig maar heeft geen beschermend effect — valse e-mail wordt niet tegengehouden. Upgrade naar p=quarantine of p=reject.")
        # quarantine en reject zijn beide goed, geen onderscheid nodig
        if not dmarc_rua:
            tips.append("DMARC heeft geen rua (rapportage-adres): je ontvangt geen meldingen over misbruik van je domein. Voeg rua=mailto:... toe.")

    # ── DNSSEC ───────────────────────────────────────────────────────────────
    if dnssec == "no":
        tips.append("DNSSEC ontbreekt: DNS-antwoorden kunnen worden vervalst (cache poisoning). Activeer DNSSEC bij je DNS-provider.")

    # ── MTA-STS ──────────────────────────────────────────────────────────────
    mta_sts      = result.get("mta_sts", "")
    mta_sts_mode = result.get("mta_sts_mode", "")
    if mta_sts == "not found":
        tips.append("MTA-STS ontbreekt: e-mail kan onversleuteld worden afgeleverd. Publiceer een MTA-STS policy om TLS af te dwingen.")
    elif mta_sts_mode and mta_sts_mode != "enforce":
        tips.append(f"MTA-STS staat op '{mta_sts_mode}': TLS wordt nog niet afgedwongen. Zet de policy op 'enforce' voor volledige bescherming.")

    return "; ".join(tips) if tips else "geen aandachtspunten"


def vraag_domein(naam: str) -> tuple[str | None, str]:
    """
    Vraagt interactief om een domein voor een item zonder website.
    Bij een niet-.nl domein wordt om bevestiging gevraagd.
    """
    print(f"\n{'─'*60}")
    print(f"  Geen website gevonden voor: {naam}")
    print(f"{'─'*60}")

    while True:
        invoer = input("  Domein (of Enter om over te slaan): ").strip().lower()

        if not invoer:
            print("  ⏭️  Overgeslagen.")
            return None, "geen domein opgegeven"

        if invoer.startswith(("http://", "https://")):
            invoer = urlparse(invoer).hostname or invoer
        if invoer.startswith("www."):
            invoer = invoer[4:]

        if not invoer.endswith(".nl"):
            bevestig = input(f"  ⚠️  '{invoer}' is geen .nl domein. Correct? (j/n of ander domein): ").strip().lower()
            if bevestig in ("j", "ja", "y", "yes"):
                pass
            elif bevestig in ("n", "nee", "no"):
                continue
            else:
                invoer = bevestig.lstrip("www.") if bevestig.startswith("www.") else bevestig

        if "." not in invoer:
            print("  ❌ Geen geldig domein. Probeer opnieuw.")
            continue

        print(f"  ✅ Domein ingesteld: {invoer}")
        return invoer, f"handmatig opgegeven: {invoer}"



# ─── Hoofd check per domein ──────────────────────────────────────────────────

def check_domain(item: dict, timeout: int = DNS_TIMEOUT) -> dict:
    naam   = item["naam"]
    url    = item["url"]
    domain = extract_domain(url)

    base = {
        "naam":          naam,
        "url_origineel": url,
        "domain":        domain or "",
        "domain_bron":   "txt" if domain else "",
        "check_date":    datetime.now().strftime("%Y-%m-%d"),
        "check_time":    datetime.now().strftime("%H:%M:%S"),
        "status":        item.get("status", ""),
        "error":         "",
    }

    # ── Geen domein → gebruik handmatig opgegeven waarde indien beschikbaar ──
    if not domain:
        gevonden, bron = item.get("_domein_handmatig", (None, ""))
        if not gevonden:
            base["error"] = "geen domein; overgeslagen"
            return base
        domain              = gevonden
        base["domain"]      = domain
        base["domain_bron"] = bron

    resolver = make_resolver(timeout)
    result   = {**base}

    # Registreer TLD
    parts = domain.split(".")
    result["tld"] = "." + parts[-1] if len(parts) >= 2 else ""

    try:
        result.update(check_a_aaaa(resolver, domain))
        mx_result = check_mx(resolver, domain)
        result.update(mx_result)
        spf_result = check_spf(resolver, domain)
        result.update(spf_result)
        result.update(check_dmarc(resolver, domain))
        result.update(check_dkim(resolver, domain))
        result.update(check_ns(resolver, domain))
        result.update(check_dnssec(resolver, domain))
        result.update(check_mta_sts(resolver, domain))
        result.update(check_bimi(resolver, domain))
        result.update(check_security_txt(domain))
        result.update(check_caa(resolver, domain))

        # ── Definitieve mail_provider: SPF heeft prioriteit boven MX ─────────
        spf_hint = spf_result.get("spf_mail_provider_hint", "")
        mx_prov  = mx_result.get("mail_provider", "")
        smg      = mx_result.get("smg", "")

        if spf_hint and spf_hint not in ("other", "onbekend"):
            definitief = spf_hint
        elif mx_prov and mx_prov not in ("other", "onbekend", ""):
            definitief = mx_prov
        elif smg:
            definitief = f"via {smg} (onbekend)"
        else:
            definitief = "onbekend"
        result["mail_provider"] = definitief

        # ── Scenario 2: geen MX → gebruik handmatig opgegeven alternatief ────
        alt = item.get("_alt_domein_handmatig")
        if not mx_result.get("mx_records") and alt:
            alt_domein, alt_bron = alt
            if alt_domein and alt_domein != domain:
                alt_mx = check_mx(resolver, alt_domein)
                if alt_mx.get("mx_records"):
                    result["mx_alternatief_domein"] = alt_domein
                    result["mx_records"]            = alt_mx["mx_records"]
                    result["mx_bron"]               = f"alternatief domein handmatig: {alt_domein}"
                    # Herhaal provider-conclusie met nieuw MX
                    if not spf_hint or spf_hint in ("other", "onbekend"):
                        result["mail_provider"] = alt_mx.get("mail_provider", "onbekend")
                else:
                    result["mx_alternatief_domein"] = alt_domein
                    result["mx_bron"]               = f"handmatig opgegeven {alt_domein} maar ook geen MX"
        elif not mx_result.get("mx_records"):
            result["mx_alternatief_domein"] = ""
            result["mx_bron"]               = ""

        # ── Tips genereren op basis van alle resultaten ───────────────────────
        result["tips"] = genereer_tips(result)

    except dns.resolver.NXDOMAIN:
        result["error"] = "domein bestaat niet (NXDOMAIN)"
    except dns.resolver.Timeout:
        result["error"] = "timeout"
    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}"

    return result


# ─── Invoer parsers ─────────────────────────────────────────────────────────


def parse_domains_txt(txt_path: str) -> list[dict]:
    """Leest een tekstbestand met één domein per regel (# = commentaar, header wordt overgeslagen)."""
    items = []
    with open(txt_path, encoding="utf-8") as f:
        for regel in f:
            regel = regel.strip()
            if not regel or regel.startswith("#") or regel.lower() == "domain":
                continue
            delen = re.split(r'\t|  +', regel, maxsplit=1)
            domein = delen[0].strip()
            naam   = delen[1].strip() if len(delen) > 1 else domein
            if domein:
                items.append({"naam": naam, "url": domein, "status": ""})

    log.info(f"📄 {len(items)} domeinen gevonden in {txt_path}")
    return items


# ─── CSV-kolommen (vaste volgorde) ──────────────────────────────────────────

COLUMNS = [
    "naam", "url_origineel", "domain", "tld", "domain_bron", "check_date", "check_time", "status", "error",
    # A / AAAA
    "a_records", "aaaa_records", "ipv6_support",
    # MX
    "mx_records", "mail_provider", "smg", "mx_alternatief_domein", "mx_bron",
    # SPF
    "spf_record", "spf_strictness", "spf_lookups", "spf_record_lengte",
    "spf_verzenders", "spf_fout", "spf_mail_provider_hint", "spf_smg",
    # DMARC
    "dmarc_record", "dmarc_policy", "dmarc_rua", "dmarc_pct",
    # DKIM
    "dkim_microsoft", "dkim_google",
    "dkim_selector1", "dkim_selector2", "dkim_google2",
    # NS / DNS
    "ns_records", "dns_provider",
    # Beveiliging
    "dnssec", "mta_sts", "mta_sts_mode", "bimi",
    # CAA
    "caa_records",
    # security.txt
    "security_txt", "security_txt_url",
    # Tips
    "tips",
]



# ─── HTML rapport generator ──────────────────────────────────────────────────

def _badge(waarde: str, goed: list, slecht: list, neutraal_kleur: str = "#6c757d") -> str:
    """Geeft een gekleurde badge terug op basis van de waarde."""
    waarde_l = (waarde or "").lower()
    # Gebruik exacte match of woordgrens om false positives te voorkomen
    # bijv. "found" mag niet matchen in "not found"
    def _match(waarde_l, termen):
        for t in termen:
            t_l = t.lower()
            if waarde_l == t_l:
                return True
            idx = waarde_l.find(t_l)
            while idx != -1:
                # Niet matchen als er "not " direct voor staat
                if idx >= 4 and waarde_l[idx-4:idx] == "not ":
                    idx = waarde_l.find(t_l, idx + 1)
                    continue
                voor = idx == 0 or not waarde_l[idx-1].isalnum()
                na   = idx + len(t_l) == len(waarde_l) or not waarde_l[idx + len(t_l)].isalnum()
                if voor and na:
                    return True
                idx = waarde_l.find(t_l, idx + 1)
        return False

    if _match(waarde_l, goed):
        kleur = "#198754"   # groen
    elif _match(waarde_l, slecht):
        kleur = "#dc3545"   # rood
    else:
        kleur = neutraal_kleur
    label = waarde or "—"
    return f'<span style="background:{kleur};color:#fff;padding:2px 10px;border-radius:12px;font-size:0.85em;font-weight:600">{label}</span>'


def _rij(label: str, waarde, goed=None, slecht=None, hint: str = "") -> str:
    """Geeft een tabelrij terug, optioneel met gekleurde badge."""
    if waarde is None or waarde == "":
        cel = '<span style="color:#aaa">—</span>'
    elif goed or slecht:
        cel = _badge(str(waarde), goed or [], slecht or [])
    else:
        cel = f'<span style="font-family:monospace;font-size:0.85em;word-break:break-all">{waarde}</span>'

    hint_html = f'<br><small style="color:#888">{hint}</small>' if hint else ""
    return f"""
        <tr>
          <td style="padding:8px 12px;color:#555;width:220px;vertical-align:top">{label}</td>
          <td style="padding:8px 12px">{cel}{hint_html}</td>
        </tr>"""


def _sectie(titel: str, rijen: str) -> str:
    return f"""
    <div style="margin-bottom:28px">
      <h2 style="font-size:1em;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
                 color:#444;border-bottom:2px solid #e9ecef;padding-bottom:6px;margin-bottom:0">{titel}</h2>
      <table style="width:100%;border-collapse:collapse">{rijen}</table>
    </div>"""


def schrijf_html_rapport(r: dict, pad: str):
    """Genereert een zelfstandig HTML-rapport voor één domein."""
    domain      = r.get("domain", "onbekend")
    check_datum = f"{r.get('check_date','')} {r.get('check_time','')}"
    tips        = r.get("tips", "")
    error       = r.get("error", "")

    # Bepaal overall score op basis van het aantal gegenereerde tips
    # (wordt na genereer_tips ingevuld — hier alvast berekenen via dezelfde logica)
    tips_lijst = [t.strip() for t in tips.split(";") if t.strip()] if tips and tips != "geen aandachtspunten" else []
    rode_items  = len(tips_lijst)
    score_kleur = "#198754" if rode_items == 0 else ("#fd7e14" if rode_items <= 2 else "#dc3545")
    score_label = "Goed" if rode_items == 0 else ("Let op" if rode_items <= 2 else "Aandacht vereist")

    # ── Secties ──────────────────────────────────────────────────────────────
    s_algemeen = _sectie("Algemeen", "".join([
        _rij("Domein",        domain),
        _rij("Gecheckt op",   check_datum),
        _rij("Status",        r.get("status") or "—"),
        _rij("Foutmelding",   error) if error else "",
        _rij("IPv4 (A)",      r.get("a_records")),
        _rij("IPv6 (AAAA)",   r.get("aaaa_records")),
        _rij("IPv6 support",  r.get("ipv6_support"),
             goed=["yes"], slecht=["no"]),
    ]))

    s_mail = _sectie("E-mail & MX", "".join([
        _rij("MX records",    r.get("mx_records")),
        _rij("Mail provider", r.get("mail_provider")),
        _rij("SMG / filter",  r.get("smg") or "geen"),
        _rij("Alt. MX domein",r.get("mx_alternatief_domein") or "—"),
    ]))

    s_spf = _sectie("SPF", "".join([
        _rij("SPF aanwezig",     "ja" if r.get("spf_record") else "nee",
             goed=["ja"], slecht=["nee"]),
        _rij("Strictheid",       r.get("spf_strictness"),
             goed=["hard fail", "soft fail", "-all", "~all", "redirect →"],
             slecht=["⚠️"],
             hint="Soft fail (~all) en hard fail (-all) zijn beide acceptabel. Bij redirect= wordt de strictheid bepaald door het doeldomein"),
        _rij("DNS lookups",      r.get("spf_lookups"),
             goed=["✓"],
             slecht=["⚠️"],
             hint="RFC 7208 staat maximaal 10 lookups toe (recursief, door alle includes heen)"),
        _rij("Recordlengte",     r.get("spf_record_lengte"),
             goed=["✓"],
             slecht=["⚠️"],
             hint="Gemeten over het samengevoegde record (meerdere TXT-strings worden door de resolver gecombineerd). Bij >255 bytes is consolidatie van includes aan te raden (RFC 7208)"),
        _rij("Effectieve verzenders", r.get("spf_verzenders"),
             hint="Providers en diensten die na uitklappen van alle includes gevonden zijn"),
        _rij("Fouten",           r.get("spf_fout") or "geen",
             slecht=["⚠️"]),
        _rij("SPF record",       r.get("spf_record")),
    ]))

    dmarc_p = r.get("dmarc_policy", "")
    s_dmarc = _sectie("DMARC", "".join([
        _rij("DMARC aanwezig", "ja" if r.get("dmarc_record") else "nee",
             goed=["ja"], slecht=["nee"]),
        _rij("Policy",        dmarc_p,
             goed=["reject", "quarantine"], slecht=["none"],
             hint="'reject' en 'quarantine' bieden beide bescherming; 'none' heeft geen effect"),
        _rij("Rapportage (rua)", r.get("dmarc_rua") or "niet ingesteld",
             goed=["mailto"], slecht=["niet ingesteld"],
             hint="rua zorgt dat je meldingen ontvangt over misbruik van je domein"),
        _rij("Percentage (pct)", r.get("dmarc_pct") or "100",
             hint="100% betekent dat de policy voor alle e-mail geldt"),
        _rij("DMARC record",  r.get("dmarc_record")),
    ]))

    ms_dkim  = r.get("dkim_microsoft", "")
    g_dkim   = r.get("dkim_google", "")
    s_dkim = _sectie("DKIM", "".join([
        _rij("Microsoft DKIM",   ms_dkim,  goed=["found"]),
        _rij("  selector1",      r.get("dkim_selector1"), goed=["found"]),
        _rij("  selector2",      r.get("dkim_selector2"), goed=["found"]),
        _rij("Google DKIM",      g_dkim,   goed=["found"]),
        _rij("  google",         r.get("dkim_google2"),   goed=["found"]),
    ]))

    s_dns = _sectie("DNS & Infrastructuur", "".join([
        _rij("NS records",    r.get("ns_records")),
        _rij("DNS provider",  r.get("dns_provider")),
        _rij("CAA records",   r.get("caa_records") or "niet ingesteld",
             hint="CAA beperkt welke CA's SSL-certificaten mogen uitgeven"),
        _rij("DNSSEC",        r.get("dnssec"),
             goed=["yes"], slecht=["no"],
             hint="DNSSEC voorkomt DNS-antwoord vervalsing"),
        _rij("MTA-STS",       r.get("mta_sts"),
             goed=["found"],
             hint="MTA-STS dwingt versleutelde e-mailbezorging af"),
        _rij("  mode",        r.get("mta_sts_mode") or "—",
             goed=["enforce"], slecht=["testing", "none", "onbekend", "niet bereikbaar"],
             hint="'enforce' dwingt TLS af; 'testing' logt alleen maar blokkeert niet; een TLS-certificaatfout op mta-sts.{domein} telt als configuratiefout"),
        _rij("BIMI",          r.get("bimi"),
             goed=["found"],
             hint="BIMI toont je logo in ondersteunde e-mailclients"),
        _rij("security.txt",  r.get("security_txt"),
             goed=["found"], slecht=["not found"],
             hint="security.txt maakt het melden van kwetsbaarheden mogelijk"),
    ]))

    # ── Tips blok ────────────────────────────────────────────────────────────
    if tips and tips != "geen aandachtspunten":
        tip_items = "".join(
            f'<li style="margin-bottom:8px">{t.strip()}</li>'
            for t in tips.split(";") if t.strip()
        )
        tips_html = f"""
    <div style="background:#fff8e1;border-left:4px solid #ffc107;padding:16px 20px;
                border-radius:0 8px 8px 0;margin-bottom:28px">
      <h2 style="font-size:1em;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
                 color:#856404;margin:0 0 12px">⚠️ Aandachtspunten</h2>
      <ul style="margin:0;padding-left:20px;color:#444;line-height:1.6">{tip_items}</ul>
    </div>"""
    else:
        tips_html = """
    <div style="background:#d1e7dd;border-left:4px solid #198754;padding:16px 20px;
                border-radius:0 8px 8px 0;margin-bottom:28px">
      <p style="margin:0;color:#0a3622;font-weight:600">✅ Geen aandachtspunten gevonden</p>
    </div>"""

    html = f"""<!DOCTYPE html>
<html lang="nl">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DNS Rapport — {domain}</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
           background: #f8f9fa; color: #212529; margin: 0; padding: 24px; }}
    .container {{ max-width: 820px; margin: 0 auto; }}
    table tr:nth-child(even) {{ background: #f8f9fa; }}
  </style>
</head>
<body>
<div class="container">

  <div style="background:#fff;border-radius:12px;padding:24px 28px;
              box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:24px;
              display:flex;justify-content:space-between;align-items:center">
    <div>
      <div style="font-size:0.75em;color:#888;text-transform:uppercase;letter-spacing:.1em">DNS Rapport</div>
      <h1 style="margin:4px 0 0;font-size:1.6em">{domain}</h1>
      <div style="color:#888;font-size:0.85em;margin-top:4px">Gecheckt op {check_datum}</div>
    </div>
    <div style="text-align:right">
      <div style="background:{score_kleur};color:#fff;padding:8px 18px;
                  border-radius:20px;font-weight:700;font-size:1em">{score_label}</div>
      <div style="color:#888;font-size:0.8em;margin-top:4px">{rode_items} aandachtspunt(en)</div>
    </div>
  </div>

  <div style="background:#fff;border-radius:12px;padding:24px 28px;
              box-shadow:0 1px 4px rgba(0,0,0,.08)">
    {tips_html}
    {s_algemeen}
    {s_mail}
    {s_spf}
    {s_dmarc}
    {s_dkim}
    {s_dns}
  </div>

  <div style="background:#e8f4f8;border-left:4px solid #0d6efd;padding:16px 20px;
              border-radius:0 8px 8px 0;margin-top:24px">
    <h3 style="margin:0 0 8px;font-size:0.95em;font-weight:700;color:#084298">
      Let op: ook inactieve domeinen verdienen bescherming
    </h3>
    <p style="margin:0;color:#444;font-size:0.88em;line-height:1.6">
      Domeinen die niet worden gebruikt voor e-mail zijn vaak een makkelijk doelwit voor spoofing,
      juist omdat er geen monitoring op zit. Voorzie ook deze domeinen van een SPF-record met
      <code>-all</code> (geen enkele server mag mailen) en een DMARC-record met minimaal
      <code>p=quarantine</code> of <code>p=reject</code>. Zo voorkom je dat aanvallers jouw
      domeinnaam misbruiken voor phishing zonder dat je het merkt.
    </p>
  </div>

  <div style="text-align:center;color:#aaa;font-size:0.78em;margin-top:20px;line-height:1.8">
    Gegenereerd door DNS Checker · {check_datum}<br>
    &copy; Sigurd Felix &mdash; De gehanteerde criteria zijn subjectief en gebaseerd op gangbare best practices.
    Aan dit rapport kunnen geen rechten worden ontleend.
  </div>
</div>
</body>
</html>"""

    with open(pad, "w", encoding="utf-8") as f:
        f.write(html)


# ─── Opstartmenu ─────────────────────────────────────────────────────────────

def _bestandskiezer(titel: str, filetypes: list) -> str | None:
    """Opent een bestandskiezer via tkinter als dat beschikbaar is."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        pad = filedialog.askopenfilename(title=titel, filetypes=filetypes)
        root.destroy()
        return pad if pad else None
    except Exception:
        return None


def opstartmenu() -> list[dict]:
    """
    Interactief opstartmenu. Wordt getoond als het script zonder argumenten
    wordt gestart (bijv. door dubbelklikken).
    Geeft een lijst met items terug.
    """
    while True:
        print()
        print("╔══════════════════════════════════════════════════╗")
        print("║                  DNS Checker                     ║")
        print("╚══════════════════════════════════════════════════╝")
        print()
        print("  Wat wil je doen?")
        print()
        print("  1. Een lijst met domeinen controleren (domains.txt)")
        print("  2. Een enkel domein controleren")
        print("  3. Over / disclaimer")
        print("  4. Afsluiten")
        print()

        keuze = input("  Keuze (1/2/3/4): ").strip()

        if keuze == "4":
            return None

        if keuze == "3":
            print()
            print("╔══════════════════════════════════════════════════╗")
            print("║              Over DNS Checker                    ║")
            print("╚══════════════════════════════════════════════════╝")
            print()
            print("  Dit script is ontwikkeld met behulp van Claude")
            print("  (Anthropic) en wordt aangeboden zoals het is.")
            print()
            print("  Er worden geen garanties gegeven over de juistheid")
            print("  van de uitkomsten of de werking van het script in")
            print("  de toekomst. DNS-standaarden en de interpretatie")
            print("  daarvan kunnen wijzigen.")
            print()
            print("  De gehanteerde beoordelingscriteria zijn subjectief.")
            print("  Raadpleeg altijd een specialist voor definitieve")
            print("  conclusies over de beveiliging van een domein.")
            print()
            print("  Vragen over de tool?")
            print("  Neem contact op via sigurd@forevertoday.nl")
            print()
            input("  Druk Enter om terug te gaan naar het menu...")
            continue

        if keuze in ("1", "2"):
            break
        print("  Voer 1, 2, 3 of 4 in.")

    items = []

    if keuze == "1":
        print()
        pad = _bestandskiezer(
            "Selecteer domains.txt",
            [("Tekstbestanden", "*.txt"), ("Alle bestanden", "*.*")]
        )
        if not pad:
            print("  Voer het pad naar domains.txt in")
            print("  (bijv. C:\\Users\\naam\\Downloads\\domains.txt)")
            pad = input("  Pad: ").strip().strip('"')

        if not pad or not os.path.exists(pad):
            print(f"  ❌ Bestand niet gevonden: {pad}")
            sys.exit(1)

        items = parse_domains_txt(pad)

    elif keuze == "2":
        print()
        while True:
            invoer = input("  Domein (bijv. mijnbedrijf.nl): ").strip().lower()
            if not invoer:
                print("  Voer een domein in.")
                continue
            if invoer.startswith(("http://", "https://")):
                invoer = urlparse(invoer).hostname or invoer
            if invoer.startswith("www."):
                invoer = invoer[4:]
            if "." not in invoer:
                print("  ❌ Geen geldig domein. Probeer opnieuw.")
                continue
            break

        items = [{"naam": invoer, "url": invoer, "status": "handmatig"}]
        print(f"  ✅ Domein ingesteld: {invoer}")

    return items


# ─── Main ────────────────────────────────────────────────────────────────────

def _maak_docs(docs_dir: str):
    """Schrijft README.txt, LICENSE.txt en DEPENDENCIES.txt naar docs_dir als ze nog niet bestaan."""

    bestanden = {
        "README.txt": """\
================================================================================
  DNS CHECKER — Handleiding
  Auteur: Sigurd Felix
  Versie: 2026
================================================================================

INHOUDSOPGAVE
  1. Wat doet DNS Checker?
  2. Vereisten en installatie
  3. Opstarten
  4. Gebruiksmodes
     4a. Opstartmenu (dubbelklik)
     4b. Enkel domein (--domain)
     4c. Lijst van domeinen (--domains)
     4d. Geavanceerde opties
  5. Uitvoer
     5a. HTML-rapport (enkelvoudig domein)
     5b. Excel-werkmap (lijst van domeinen)
  6. Uitleg van de checks
     6a. SPF
     6b. DMARC
     6c. DKIM
     6d. DNSSEC
     6e. MTA-STS
     6f. BIMI
     6g. security.txt
     6h. CAA
  7. Kleurcodes
  8. Disclaimer


--------------------------------------------------------------------------------
1. WAT DOET DNS CHECKER?
--------------------------------------------------------------------------------

DNS Checker analyseert de DNS-configuratie van domeinen op het gebied van
e-mailbeveiliging en infrastructuur. Het script controleert per domein een
reeks records (SPF, DMARC, DKIM, DNSSEC, MTA-STS en meer) en geeft per
bevinding aan of de configuratie goed, onvolledig of afwezig is.

Het resultaat is een HTML-rapport (voor één domein) of een gekleurde
Excel-werkmap (voor een lijst van domeinen), waarbij een groene cel aangeeft
dat de configuratie in orde is, een rode cel een aandachtspunt markeert en een
grijze cel aangeeft dat een instelling niet van toepassing of niet gevonden is.

DNS Checker is bedoeld als hulpmiddel bij het snel in kaart brengen van de
e-mailbeveiligingsstatus van een groot aantal domeinen, bijvoorbeeld bij een
sectoranalyse of een interne audit.


--------------------------------------------------------------------------------
2. VEREISTEN EN INSTALLATIE
--------------------------------------------------------------------------------

Vereisten:
  - Python 3.10 of hoger
  - Internetverbinding (voor DNS-lookups en HTTP-checks)

Benodigde Python-pakketten:
  - dnspython
  - requests
  - openpyxl

Bij het opstarten controleert DNS Checker automatisch of de benodigde pakketten
aanwezig zijn. Als dat niet het geval is, wordt gevraagd of ze automatisch
geïnstalleerd mogen worden via pip.

Handmatig installeren:
  pip install dnspython requests openpyxl


--------------------------------------------------------------------------------
3. OPSTARTEN
--------------------------------------------------------------------------------

Dubbelklik op dns_checker.py (Windows: rechtermuisknop → Openen met → Python),
of start het script vanaf de command prompt:

  python dns_checker.py

Als er een domains.txt staat in dezelfde map als het script, start DNS Checker
automatisch met die lijst. Anders verschijnt het opstartmenu.

Bij de eerste opstart worden automatisch de mappen docs/ en reports/ aangemaakt
naast het script, en worden deze documentatiebestanden gegenereerd in docs/.


--------------------------------------------------------------------------------
4. GEBRUIKSMODES
--------------------------------------------------------------------------------

4a. OPSTARTMENU (dubbelklik / geen argumenten)
-----------------------------------------------
Als het script zonder argumenten wordt gestart, verschijnt het opstartmenu:

  1. Een lijst met domeinen controleren (domains.txt)
  2. Een enkel domein controleren
  3. Over / disclaimer
  4. Afsluiten

Keuze 1: Er wordt een bestandskiezer geopend waarmee je een domains.txt kunt
selecteren. Als tkinter niet beschikbaar is, kun je het pad handmatig invoeren.

Keuze 2: Voer een domein in (bijv. mijnbedrijf.nl of https://www.mijnbedrijf.nl).
DNS Checker haalt automatisch het domein op uit een volledige URL. Het resultaat
wordt opgeslagen als HTML-rapport in de map reports/.

Keuze 3: Toont informatie over het script, de auteur en de disclaimer.

Na elke uitvoering keer je terug naar het menu. Kies 4 om af te sluiten.


4b. ENKEL DOMEIN (--domain)
----------------------------
Controleer één domein direct vanuit de command prompt:

  python dns_checker.py --domain mijnbedrijf.nl
  python dns_checker.py --domain https://www.mijnbedrijf.nl

Het resultaat wordt opgeslagen in reports/:
  reports/dns_rapport_mijnbedrijf.nl_20260308_143022.html


4c. LIJST VAN DOMEINEN (--domains)
------------------------------------
Verwerk een tekstbestand met meerdere domeinen:

  python dns_checker.py --domains domains.txt

Het tekstbestand bevat één domein per regel. Commentaarregels beginnen met #.
Een eerste headerregel "domain" wordt automatisch overgeslagen. Voorbeeld:

  # DNS controle - maart 2026
  domain
  mijnbedrijf.nl
  anderebedrijf.nl
  derdebedrijf.nl

Het resultaat wordt opgeslagen in reports/:
  reports/dns_results_20260308_143022.xlsx


4d. GEAVANCEERDE OPTIES
------------------------
  --domain   <domein>    Enkel domein controleren
  --domains  <bestand>   Pad naar domains.txt
  --workers  <n>         Aantal parallelle threads (standaard: 5)
  --timeout  <n>         DNS timeout in seconden (standaard: 5)
  --no-input             Geen interactieve vragen (voor automatisering)

Voorbeelden:
  python dns_checker.py --domains domains.txt --workers 10
  python dns_checker.py --domains domains.txt --timeout 3 --no-input
  python dns_checker.py --domain mijnbedrijf.nl


--------------------------------------------------------------------------------
5. UITVOER
--------------------------------------------------------------------------------

Alle uitvoerbestanden worden opgeslagen in de map reports/ naast het script.

5a. HTML-RAPPORT (enkelvoudig domein)
---------------------------------------
Het HTML-rapport bevat een overzicht van alle DNS-checks voor één domein,
inclusief een overall score (Goed / Let op / Aandacht vereist), een blok met
aandachtspunten en gedetailleerde secties per onderwerp.

Het rapport is een zelfstandig HTML-bestand dat in elke webbrowser kan worden
geopend en per e-mail kan worden doorgestuurd.

Secties in het rapport:
  - Algemeen       : domein, checkdatum, IPv4/IPv6, NS-provider
  - E-mail & MX    : MX-records, mailprovider, SMG-detectie
  - SPF            : record, strictheid, lookups, lengte, verzenders, fouten
  - DMARC          : record, policy, rapportage-adres (rua), percentage
  - DKIM           : Microsoft (selector1/selector2) en Google (google/google2)
  - DNS & infra    : DNSSEC, MTA-STS + mode, BIMI, CAA, security.txt

Bestandsnaam: reports/dns_rapport_{domein}_{timestamp}.html


5b. EXCEL-WERKMAP (lijst van domeinen)
----------------------------------------
De Excel-werkmap bevat één tabblad met alle domeinen als rijen. Elke kolom
stelt één DNS-check voor en krijgt een achtergrondkleur op basis van de uitkomst.

Kolommen: Naam, Domein, Mail provider, SMG, MX records, SPF (record +
strictheid + lookups + lengte + verzenders + fouten), DMARC (policy + rua),
DKIM (Microsoft + Google), DNSSEC, MTA-STS + mode, BIMI, security.txt,
NS provider, Check datum.

De eerste rij en kolom A zijn bevroren voor eenvoudig navigeren.

Bestandsnaam: reports/dns_results_{timestamp}.xlsx


--------------------------------------------------------------------------------
6. UITLEG VAN DE CHECKS
--------------------------------------------------------------------------------

6a. SPF (Sender Policy Framework)
------------------------------------
SPF legt vast welke mailservers e-mail mogen versturen namens een domein.

  Strictheid:
    -all (hard fail)  → niet-geautoriseerde mail wordt geweigerd       ✓
    ~all (soft fail)  → niet-geautoriseerde mail wordt gemarkeerd       ✓
    +all              → IEDEREEN mag mailen — biedt geen bescherming    ✗
    ?all              → neutraal, ontvangende server doet niets         ✗

  DNS lookups : RFC 7208 staat maximaal 10 toe (inclusief geneste includes)
  Recordlengte: gemeten over het samengevoegde record; bij >255 bytes is consolidatie aan te raden (RFC 7208)
  Verzenders  : providers die na uitklappen van alle includes gevonden zijn

Meerdere SPF-records op één domein is een fout (RFC 7208 staat er maar één toe).

6b. DMARC
----------
DMARC bouwt voort op SPF en DKIM en bepaalt wat er gebeurt met mail die niet
voldoet aan de authenticatiecontroles.

  p=none        → geen actie, alleen rapportage (geen bescherming)
  p=quarantine  → verdachte mail naar spam verplaatsen
  p=reject      → verdachte mail weigeren

  rua: rapportage-adres voor geaggregeerde rapporten. Zonder rua ontvang
       je geen meldingen over misbruik van je domein.

6c. DKIM (DomainKeys Identified Mail)
---------------------------------------
DKIM voegt een cryptografische handtekening toe aan uitgaande e-mail.
DNS Checker controleert de meest gangbare selectors:
  Microsoft 365 : selector1 en selector2
  Google        : google en google2

Niet-gevonden selectors zijn grijs — andere selectors zijn mogelijk in gebruik.

6d. DNSSEC
-----------
DNSSEC voegt digitale handtekeningen toe aan DNS-antwoorden en beschermt
tegen DNS-vervalsing (cache poisoning).

6e. MTA-STS
------------
MTA-STS dwingt verzendende mailservers om TLS te gebruiken bij aflevering.

  mode: testing  → actief maar TLS wordt nog niet afgedwongen
  mode: enforce  → TLS volledig afgedwongen

6f. BIMI
---------
BIMI maakt het mogelijk een merklogo te tonen in ondersteunde e-mailclients.
Vereist DMARC met p=quarantine of p=reject.

6g. security.txt
-----------------
Gestandaardiseerd bestand (RFC 9116) waarmee organisaties aangeven hoe
beveiligingsonderzoekers kwetsbaarheden kunnen melden.

6h. CAA
--------
CAA-records beperken welke certificaatautoriteiten SSL/TLS-certificaten
mogen uitgeven voor een domein.


--------------------------------------------------------------------------------
7. KLEURCODES
--------------------------------------------------------------------------------

  GROEN  : configuratie is correct en voldoet aan best practices
  ROOD   : aandacht vereist — ontbreekt of is onjuist geconfigureerd
  GRIJS  : niet van toepassing, niet gevonden of neutraal
  WIT    : puur informatief veld zonder kwaliteitsoordeel


--------------------------------------------------------------------------------
8. DISCLAIMER
--------------------------------------------------------------------------------

Dit script is ontwikkeld met behulp van Claude (Anthropic) en wordt aangeboden
zoals het is. Er worden geen garanties gegeven over de juistheid van de
uitkomsten of de werking in de toekomst. DNS-standaarden kunnen wijzigen.

De gehanteerde beoordelingscriteria zijn subjectief. Raadpleeg altijd een
specialist voor definitieve conclusies. Aan de uitkomsten kunnen geen rechten
worden ontleend.

Vragen? Neem contact op via sigurd@forevertoday.nl

================================================================================
""",

        "LICENSE.txt": """\
                                 Apache License
                           Version 2.0, January 2004
                        http://www.apache.org/licenses/

   TERMS AND CONDITIONS FOR USE, REPRODUCTION, AND DISTRIBUTION

   1. Definitions.

      "License" shall mean the terms and conditions for use, reproduction,
      and distribution as defined by Sections 1 through 9 of this document.

      "Licensor" shall mean the copyright owner or entity authorized by
      the copyright owner that is granting the License.

      "Legal Entity" shall mean the union of the acting entity and all
      other entities that control, are controlled by, or are under common
      control with that entity. For the purposes of this definition,
      "control" means (i) the power, direct or indirect, to cause the
      direction or management of such entity, whether by contract or
      otherwise, or (ii) ownership of fifty percent (50%) or more of the
      outstanding shares, or (iii) beneficial ownership of such entity.

      "You" (or "Your") shall mean an individual or Legal Entity
      exercising permissions granted by this License.

      "Source" form shall mean the preferred form for making modifications,
      including but not limited to software source code, documentation
      source, and configuration files.

      "Object" form shall mean any form resulting from mechanical
      transformation or translation of a Source form, including but
      not limited to compiled object code, generated documentation,
      and conversions to other media types.

      "Work" shall mean the work of authorship, whether in Source or
      Object form, made available under the License, as indicated by a
      copyright notice that is included in or attached to the work
      (an example is provided in the Appendix below).

      "Derivative Works" shall mean any work, whether in Source or Object
      form, that is based on (or derived from) the Work and for which the
      editorial revisions, annotations, elaborations, or other modifications
      represent, as a whole, an original work of authorship. For the purposes
      of this License, Derivative Works shall not include works that remain
      separable from, or merely link (or bind by name) to the interfaces of,
      the Work and Derivative Works thereof.

      "Contribution" shall mean any work of authorship, including
      the original version of the Work and any modifications or additions
      to that Work or Derivative Works thereof, that is intentionally
      submitted to Licensor for inclusion in the Work by the copyright owner
      or by an individual or Legal Entity authorized to submit on behalf of
      the copyright owner. For the purposes of this definition, "submitted"
      means any form of electronic, verbal, or written communication sent
      to the Licensor or its representatives, including but not limited to
      communication on electronic mailing lists, source code control systems,
      and issue tracking systems that are managed by, or on behalf of, the
      Licensor for the purpose of discussing and improving the Work, but
      excluding communication that is conspicuously marked or otherwise
      designated in writing by the copyright owner as "Not a Contribution."

      "Contributor" shall mean Licensor and any individual or Legal Entity
      on behalf of whom a Contribution has been received by Licensor and
      subsequently incorporated within the Work.

   2. Grant of Copyright License. Subject to the terms and conditions of
      this License, each Contributor hereby grants to You a perpetual,
      worldwide, non-exclusive, no-charge, royalty-free, irrevocable
      copyright license to reproduce, prepare Derivative Works of,
      publicly display, publicly perform, sublicense, and distribute the
      Work and such Derivative Works in Source or Object form.

   3. Grant of Patent License. Subject to the terms and conditions of
      this License, each Contributor hereby grants to You a perpetual,
      worldwide, non-exclusive, no-charge, royalty-free, irrevocable
      (except as stated in this section) patent license to make, have made,
      use, offer to sell, sell, import, and otherwise transfer the Work,
      where such license applies only to those patent claims licensable
      by such Contributor that are necessarily infringed by their
      Contribution(s) alone or by combination of their Contribution(s)
      with the Work to which such Contribution(s) was submitted. If You
      institute patent litigation against any entity (including a
      cross-claim or counterclaim in a lawsuit) alleging that the Work
      or a Contribution incorporated within the Work constitutes direct
      or contributory patent infringement, then any patent licenses
      granted to You under this License for that Work shall terminate
      as of the date such litigation is filed.

   4. Redistribution. You may reproduce and distribute copies of the
      Work or Derivative Works thereof in any medium, with or without
      modifications, and in Source or Object form, provided that You
      meet the following conditions:

      (a) You must give any other recipients of the Work or
          Derivative Works a copy of this License; and

      (b) You must cause any modified files to carry prominent notices
          stating that You changed the files; and

      (c) You must retain, in the Source form of any Derivative Works
          that You distribute, all copyright, patent, trademark, and
          attribution notices from the Source form of the Work,
          excluding those notices that do not pertain to any part of
          the Derivative Works; and

      (d) If the Work includes a "NOTICE" text file as part of its
          distribution, then any Derivative Works that You distribute must
          include a readable copy of the attribution notices contained
          within such NOTICE file, excluding those notices that do not
          pertain to any part of the Derivative Works, in at least one
          of the following places: within a NOTICE text file distributed
          as part of the Derivative Works; within the Source form or
          documentation, if provided along with the Derivative Works; or,
          within a display generated by the Derivative Works, if and
          wherever such third-party notices normally appear. The contents
          of the NOTICE file are for informational purposes only and
          do not modify the License. You may add Your own attribution
          notices within Derivative Works that You distribute, alongside
          or as an addendum to the NOTICE text from the Work, provided
          that such additional attribution notices cannot be construed
          as modifying the License.

      You may add Your own copyright statement to Your modifications and
      may provide additional or different license terms and conditions
      for use, reproduction, or distribution of Your modifications, or
      for any such Derivative Works as a whole, provided Your use,
      reproduction, and distribution of the Work otherwise complies with
      the conditions stated in this License.

   5. Submission of Contributions. Unless You explicitly state otherwise,
      any Contribution intentionally submitted for inclusion in the Work
      by You to the Licensor shall be under the terms and conditions of
      this License, without any additional terms or conditions.
      Notwithstanding the above, nothing herein shall supersede or modify
      the terms of any separate license agreement you may have executed
      with Licensor regarding such Contributions.

   6. Trademarks. This License does not grant permission to use the trade
      names, trademarks, service marks, or product names of the Licensor,
      except as required for reasonable and customary use in describing the
      origin of the Work and reproducing the content of the NOTICE file.

   7. Disclaimer of Warranty. Unless required by applicable law or
      agreed to in writing, Licensor provides the Work (and each
      Contributor provides its Contributions) on an "AS IS" BASIS,
      WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
      implied, including, without limitation, any warranties or conditions
      of TITLE, NON-INFRINGEMENT, MERCHANTABILITY, or FITNESS FOR A
      PARTICULAR PURPOSE. You are solely responsible for determining the
      appropriateness of using or redistributing the Work and assume any
      risks associated with Your exercise of permissions under this License.

   8. Limitation of Liability. In no event and under no legal theory,
      whether in tort (including negligence), contract, or otherwise,
      unless required by applicable law (such as deliberate and grossly
      negligent acts) or agreed to in writing, shall any Contributor be
      liable to You for damages, including any direct, indirect, special,
      incidental, or consequential damages of any character arising as a
      result of this License or out of the use or inability to use the
      Work (including but not limited to damages for loss of goodwill,
      work stoppage, computer failure or malfunction, or any and all
      other commercial damages or losses), even if such Contributor
      has been advised of the possibility of such damages.

   9. Accepting Warranty or Additional Liability. While redistributing
      the Work or Derivative Works thereof, You may choose to offer,
      and charge a fee for, acceptance of support, warranty, indemnity,
      or other liability obligations and/or rights consistent with this
      License. However, in accepting such obligations, You may act only
      on your own behalf and on your sole responsibility, not on behalf
      of any other Contributor, and only if You agree to indemnify,
      defend, and hold each Contributor harmless for any liability
      incurred by, or claims asserted against, such Contributor by reason
      of your accepting any such warranty or additional liability.

   END OF TERMS AND CONDITIONS

   Copyright 2026 Sigurd Felix

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
""",

        "DEPENDENCIES.txt": """\
================================================================================
  DNS CHECKER — Afhankelijkheden (Dependencies)
  Bijgewerkt: maart 2026
================================================================================

--------------------------------------------------------------------------------
EXTERNE PAKKETTEN (installatie via pip)
--------------------------------------------------------------------------------

1. dnspython
   Beschrijving : DNS-bibliotheek voor Python. Gebruikt voor alle DNS-lookups.
   Eigenaar     : Bob Halley / dnspython-gemeenschap
   Pip-naam     : dnspython
   Website      : https://www.dnspython.org
   Repository   : https://github.com/rthalley/dnspython
   Licentie     : ISC License (functioneel equivalent aan MIT)

2. requests
   Beschrijving : HTTP-bibliotheek. Gebruikt voor security.txt en MTA-STS.
   Eigenaar     : Kenneth Reitz / Python Software Foundation
   Pip-naam     : requests
   Website      : https://requests.readthedocs.io
   Repository   : https://github.com/psf/requests
   Licentie     : Apache License 2.0

3. openpyxl
   Beschrijving : Excel-bibliotheek. Gebruikt voor de gekleurde werkmap.
   Eigenaar     : Eric Gazoni / Charlie Clark
   Pip-naam     : openpyxl
   Website      : https://openpyxl.readthedocs.io
   Repository   : https://foss.heptapod.net/openpyxl/openpyxl
   Licentie     : MIT License

--------------------------------------------------------------------------------
PYTHON STANDAARDBIBLIOTHEEK (geen installatie vereist)
--------------------------------------------------------------------------------

argparse, base64, concurrent.futures, csv, datetime, logging, os, re,
struct, sys, tkinter, urllib.parse

Licentie: Python Software Foundation License (PSF)
Zie: https://docs.python.org/3/license.html

--------------------------------------------------------------------------------
MODELONDERSTEUNING
--------------------------------------------------------------------------------

Ontwikkeld met behulp van Claude (Anthropic) — https://www.anthropic.com

--------------------------------------------------------------------------------
LICENTIESAMENVATTING
--------------------------------------------------------------------------------

  Pakket        Licentie       Commercieel gebruik  Bronvermelding vereist
  ──────────    ─────────────  ───────────────────  ──────────────────────
  dns_checker   Apache 2.0     Ja                   Ja (bij wijzigingen)
  dnspython     ISC            Ja                   Ja
  requests      Apache 2.0     Ja                   Ja (bij wijzigingen)
  openpyxl      MIT            Ja                   Ja
  Python stdlib PSF            Ja                   Ja

================================================================================
""",

        "VERSION_1.0.txt": """\
================================================================================
  DNS CHECKER — Versie 1.0
  Releasedatum: 8 maart 2026
  Auteur: Sigurd Felix
================================================================================

EERSTE RELEASE

Functionaliteit:
  - DNS-checks voor SPF, DMARC, DKIM, DNSSEC, MTA-STS, BIMI, CAA en
    security.txt
  - SPF-analyse: recursief uitklappen van includes, lookup-teller (RFC max 10),
    recordlengte-check (RFC max 255 bytes), detectie van effectieve verzenders,
    detectie van dubbele SPF-records
  - DMARC-analyse: policy, rua-adres, percentage
  - DKIM: controle op Microsoft (selector1/selector2) en Google (google/google2)
    selectors, inclusief CNAME-opvolging voor Microsoft 365
  - MTA-STS: aanwezigheid én policy-mode (enforce / testing / none)
  - Detectie van mailprovider (Microsoft 365, Google Workspace) en SMG
    (Mimecast, Proofpoint, Hornetsecurity, Barracuda e.a.)
  - Uitvoer als zelfstandig HTML-rapport (enkelvoudig domein) of gekleurde
    Excel-werkmap (lijst van domeinen)
  - Opstartmenu met opties voor lijst, enkel domein, disclaimer en afsluiten
  - Opdrachtregelargumenten: --domain, --domains, --workers, --timeout,
    --no-input
  - Automatische aanmaak van mappen reports/ en docs/ bij eerste opstart
  - Automatische generatie van README.txt, LICENSE.txt, DEPENDENCIES.txt en
    dit versiebestand bij eerste opstart
  - Dependency-check met optie tot automatische installatie via pip
  - Uitgebracht onder de Apache License 2.0 (Copyright 2026 Sigurd Felix)
  - Disclaimer en auteursinformatie in HTML-rapporten en opstartmenu

================================================================================
""",
    }

    for bestandsnaam, inhoud in bestanden.items():
        pad = os.path.join(docs_dir, bestandsnaam)
        if not os.path.exists(pad):
            with open(pad, "w", encoding="utf-8") as f:
                f.write(inhoud)
            log.info(f"📄 {bestandsnaam} aangemaakt in docs/")


def main():
    parser = argparse.ArgumentParser(description="DNS Checker")
    parser.add_argument("--domains", default=None,       help="Tekstbestand met domeinen (één per regel)")
    parser.add_argument("--domain",  default=None,       help="Enkel domein om te controleren (bijv. mijnbedrijf.nl)")
    parser.add_argument("--workers",  type=int, default=MAX_WORKERS, help="Aantal parallelle threads")
    parser.add_argument("--timeout",  type=int, default=DNS_TIMEOUT, help="DNS timeout in seconden")
    parser.add_argument("--http-timeout", type=int, default=HTTP_TIMEOUT, help="HTTP timeout in seconden (security.txt, MTA-STS)")
    parser.add_argument("--insecure", action="store_true", help="TLS-waarschuwingen onderdrukken bij HTTP-checks")
    parser.add_argument("--no-input", action="store_true", help="Geen interactieve vragen")
    args = parser.parse_args()

    # ── Mappen aanmaken naast het script ────────────────────────────────────
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, "reports")
    docs_dir    = os.path.join(script_dir, "docs")

    for d in (reports_dir, docs_dir):
        os.makedirs(d, exist_ok=True)

    # Docs aanmaken als ze nog niet bestaan
    _maak_docs(docs_dir)

    if args.insecure:
        import dns_checker as _self  # noqa — patch eigen module globals
        requests.packages.urllib3.disable_warnings()
        log.warning("⚠️  --insecure: TLS-certificaatwaarschuwingen worden onderdrukt")

    # Pas module-globals aan zodat alle functies de CLI-waarden gebruiken
    import sys as _sys
    _mod = _sys.modules[__name__]
    _mod.HTTP_TIMEOUT = args.http_timeout
    _mod.INSECURE     = args.insecure

    via_commandline = any([args.domains, args.domain, os.path.exists("domains.txt")])

    while True:
        if args.domain:
            domein = args.domain.lower().strip()
            if domein.startswith(("http://", "https://")):
                domein = urlparse(domein).hostname or domein
            if domein.startswith("www."):
                domein = domein[4:]
            items = [{"naam": domein, "url": domein, "status": "handmatig"}]
        elif args.domains:
            items = parse_domains_txt(args.domains)
        elif os.path.exists("domains.txt"):
            items = parse_domains_txt("domains.txt")
            log.info("📄 Automatisch domains.txt gevonden")
        else:
            items = opstartmenu()

        if items is None:
            # Gebruiker koos "Afsluiten" in het menu
            print("\n  Tot ziens!\n")
            break

        _voer_checks_uit(items, args, reports_dir)

        # Na uitvoering: terug naar menu of stoppen
        if via_commandline or args.no_input:
            break
        print()
        input("  Druk Enter om terug te gaan naar het menu...")



# ─── Excel rapport generator ────────────────────────────────────────────────

def schrijf_excel_rapport(resultaten: list[dict], pad: str):
    """
    Schrijft een gekleurde Excel-werkmap met één rij per domein.
    Kolommen krijgen een groene/rode/grijze achtergrond op basis van de waarde.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Kleuren
    GROEN   = PatternFill("solid", start_color="C6EFCE")   # lichtgroen
    ROOD    = PatternFill("solid", start_color="FFC7CE")    # lichtrood
    GRIJS   = PatternFill("solid", start_color="EDEDED")    # lichtgrijs
    HOOFD   = PatternFill("solid", start_color="2F4F6F")    # donkerblauw header
    WIT     = PatternFill("solid", start_color="FFFFFF")

    RAND = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # Kolommen die we tonen, met label en kleurlogica
    # kleur: functie(waarde, rij) → "groen" | "rood" | "grijs" | "neutraal"
    def _spf_kleur(v, r):
        if not r.get("spf_record"):           return "rood"
        if "⚠️" in str(v):                    return "rood"
        if "✓" in str(v):                     return "groen"
        strict = r.get("spf_strictness", "")
        if "⚠️" in strict:                    return "rood"
        return "groen"

    def _dmarc_policy_kleur(v, r):
        if not v or v == "none":              return "rood"
        if v in ("reject", "quarantine"):     return "groen"
        return "grijs"

    def _found_kleur(v, r):
        vs = str(v or "").lower()
        if vs == "found":                     return "groen"
        if vs == "not found":                 return "grijs"
        return "neutraal"

    def _found_rood_kleur(v, r):
        vs = str(v or "").lower()
        if vs == "found":                     return "groen"
        if vs == "not found":                 return "rood"
        return "neutraal"

    def _mta_mode_kleur(v, r):
        if r.get("mta_sts") != "found":       return "grijs"
        vs = str(v or "").lower()
        if vs == "enforce":                    return "groen"
        if "⚠️" in str(v or ""):              return "rood"
        return "rood"

    def _dnssec_kleur(v, r):
        if str(v or "").lower() == "yes":     return "groen"
        if str(v or "").lower() == "no":      return "rood"
        return "grijs"

    def _rua_kleur(v, r):
        if not r.get("dmarc_record"):         return "grijs"
        if v and "mailto" in str(v).lower():  return "groen"
        return "rood"

    def _spf_lookups_kleur(v, r):
        if not v:                             return "grijs"
        if "⚠️" in str(v):                    return "rood"
        return "groen"

    def _spf_lengte_kleur(v, r):
        if not v:                             return "grijs"
        if "⚠️" in str(v):                    return "rood"
        return "groen"

    EXCEL_KOLOMMEN = [
        # label,               sleutel,             kleur_fn
        ("Naam",               "naam",               None),
        ("Domein",             "domain",             None),
        ("Mail provider",      "mail_provider",      None),
        ("SMG",                "smg",                None),
        ("MX records",         "mx_records",         None),
        ("SPF",                "spf_record",         lambda v,r: "groen" if v else "rood"),
        ("SPF strictheid",     "spf_strictness",     lambda v,r: "rood" if "⚠️" in str(v or "") else "groen" if v else "grijs"),
        ("SPF lookups",        "spf_lookups",        _spf_lookups_kleur),
        ("SPF lengte",         "spf_record_lengte",  _spf_lengte_kleur),
        ("SPF verzenders",     "spf_verzenders",     None),
        ("SPF fouten",         "spf_fout",           lambda v,r: "rood" if v else "groen"),
        ("DMARC policy",       "dmarc_policy",       _dmarc_policy_kleur),
        ("DMARC rua",          "dmarc_rua",          _rua_kleur),
        ("DKIM Microsoft",     "dkim_microsoft",     _found_kleur),
        ("  selector1",        "dkim_selector1",     _found_kleur),
        ("  selector2",        "dkim_selector2",     _found_kleur),
        ("DKIM Google",        "dkim_google",        _found_kleur),
        ("  google",           "dkim_google2",       _found_kleur),
        ("DNSSEC",             "dnssec",             _dnssec_kleur),
        ("MTA-STS",            "mta_sts",            _found_kleur),
        ("MTA-STS mode",       "mta_sts_mode",       _mta_mode_kleur),
        ("BIMI",               "bimi",               _found_kleur),
        ("security.txt",       "security_txt",       _found_rood_kleur),
        ("NS provider",        "dns_provider",       None),
        ("Check datum",        "check_date",         None),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Resultaten"
    ws.freeze_panes = "B2"  # freeze rij 1 en kolom A

    # Header rij
    for col_idx, (label, _, _) in enumerate(EXCEL_KOLOMMEN, 1):
        cel = ws.cell(row=1, column=col_idx, value=label)
        cel.fill   = HOOFD
        cel.font   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = RAND

    ws.row_dimensions[1].height = 32

    # Dataregels
    KLEUR_MAP = {"groen": GROEN, "rood": ROOD, "grijs": GRIJS, "neutraal": WIT}

    for rij_idx, record in enumerate(resultaten, 2):
        for col_idx, (_, sleutel, kleur_fn) in enumerate(EXCEL_KOLOMMEN, 1):
            waarde = record.get(sleutel, "") or ""
            cel = ws.cell(row=rij_idx, column=col_idx, value=waarde)
            cel.font      = Font(name="Arial", size=9)
            cel.alignment = Alignment(vertical="center", wrap_text=False)
            cel.border    = RAND

            if kleur_fn:
                kleur_naam = kleur_fn(waarde, record)
                cel.fill = KLEUR_MAP.get(kleur_naam, WIT)

        # Afwisselend lichte streep op rijen zonder kleur (kolom A en tekst-kolommen)
        ws.row_dimensions[rij_idx].height = 15

    # Kolombreedtes
    BREEDTES = {
        1: 28,   # Naam
        2: 22,   # Domein
        3: 18,   # Mail provider
        4: 14,   # SMG
        5: 28,   # MX records
        6: 36,   # SPF record
        7: 18,   # SPF strictheid
        8: 12,   # SPF lookups
        9: 12,   # SPF lengte
        10: 28,  # SPF verzenders
        11: 22,  # SPF fouten
        12: 14,  # DMARC policy
        13: 28,  # DMARC rua
        14: 14,  # DKIM Microsoft
        15: 12,  # selector1
        16: 12,  # selector2
        17: 12,  # DKIM Google
        18: 12,  # google
        19: 10,  # DNSSEC
        20: 10,  # MTA-STS
        21: 14,  # MTA-STS mode
        22: 10,  # BIMI
        23: 12,  # security.txt
        24: 18,  # NS provider
        25: 12,  # Check datum
    }
    for col_idx, breedte in BREEDTES.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = breedte

    wb.save(pad)


def _voer_checks_uit(items: list, args, reports_dir: str = "."):
    """Voert alle DNS-checks uit op de opgegeven items en schrijft de uitvoer."""

    met_website    = [k for k in items if k["url"].strip()]
    zonder_website = [k for k in items if not k["url"].strip()]
    log.info(f"🌐 {len(met_website)} met website, {len(zonder_website)} zonder website")

    # ── Fase 1: interactief domeinen ophalen voor items zonder website ─────
    if zonder_website and not args.no_input:
        print(f"\n{'═'*60}")
        print(f"  {len(zonder_website)} items hebben geen website.")
        print(f"  Vul per item het domein in, of druk Enter om over te slaan.")
        print(f"{'═'*60}")

        for k in zonder_website:
            domein, bron = vraag_domein(k["naam"])
            k["_domein_handmatig"] = (domein, bron)

        print(f"\n{'═'*60}")
        print(f"  Domein-invoer klaar. Start DNS-checks...")
        print(f"{'═'*60}\n")
    else:
        for k in zonder_website:
            k["_domein_handmatig"] = (None, "geen invoer (--no-input)")

    # ── Fase 2: parallel DNS-checks uitvoeren ────────────────────────────────
    ts              = datetime.now().strftime("%Y%m%d_%H%M%S")
    resultaten_dict: dict[str, dict] = {}
    totaal          = len(items)

    log.info(f"🔍 Start DNS-checks met {args.workers} workers...")

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(check_domain, k, args.timeout): k for k in items}
        for i, future in enumerate(as_completed(futures), 1):
            result = future.result()
            resultaten_dict[result["naam"]] = result
            domain = result.get("domain") or result.get("naam", "?")
            pct    = round(i / totaal * 100)
            log.info(f"  [{i:>4}/{totaal}] {pct:>3}%  {domain}")

    # ── Fase 3: interactief alternatief MX-domein vragen bij ontbrekend MX ───
    if not args.no_input:
        zonder_mx = [
            r for r in resultaten_dict.values()
            if not r.get("mx_records") and r.get("domain")
        ]
        if zonder_mx:
            print(f"\n{'═'*60}")
            print(f"  {len(zonder_mx)} items hebben geen MX-record.")
            print(f"  Mogelijk gebruiken ze een ander domein voor e-mail.")
            print(f"{'═'*60}")

            resolver = make_resolver(args.timeout)
            for result in zonder_mx:
                naam   = result["naam"]
                domain = result["domain"]
                print(f"\n  Item   : {naam}")
                print(f"  Domein : {domain} (geen MX gevonden)")
                invoer = input("  Alt. mail-domein (of Enter om over te slaan): ").strip().lower()

                if not invoer:
                    continue
                if invoer.startswith("www."):
                    invoer = invoer[4:]

                if not invoer.endswith(".nl"):
                    bevestig = input(f"  ⚠️  '{invoer}' is geen .nl domein. Correct? (j/n): ").strip().lower()
                    if bevestig not in ("j", "ja", "y", "yes"):
                        continue

                alt_mx = check_mx(resolver, invoer)
                if alt_mx.get("mx_records"):
                    result["mx_alternatief_domein"] = invoer
                    result["mx_records"]            = alt_mx["mx_records"]
                    result["mail_provider"]         = alt_mx["mail_provider"]
                    result["mx_bron"]               = f"alternatief domein handmatig: {invoer}"
                    print(f"  ✅ MX gevonden: {alt_mx['mx_records']} ({alt_mx['mail_provider']})")
                else:
                    print(f"  ❌ Geen MX op {invoer} gevonden.")

    # ── Fase 4: uitvoer — Excel voor lijsten, HTML voor enkel domein ──────────
    resultaten   = sorted(resultaten_dict.values(), key=lambda r: r.get("naam", "").lower())
    enkelvoudig  = (len(resultaten) == 1)

    if enkelvoudig:
        r         = resultaten[0]
        html_path = os.path.join(reports_dir, f"dns_rapport_{r.get('domain', 'onbekend')}_{ts}.html")
        schrijf_html_rapport(r, html_path)
        log.info(f"\n✅ Rapport opgeslagen: {html_path}")
    else:
        xlsx_path = os.path.join(reports_dir, f"dns_results_{ts}.xlsx")
        schrijf_excel_rapport(resultaten, xlsx_path)
        log.info(f"\n✅ Resultaten opgeslagen: {xlsx_path}")
        log.info(f"   Totaal gecheckt: {len(resultaten)} domeinen")

        ms_count = sum(1 for r in resultaten if r.get("mail_provider") == "Microsoft 365")
        g_count  = sum(1 for r in resultaten if r.get("mail_provider") == "Google Workspace")
        dmarc_ok = sum(1 for r in resultaten if r.get("dmarc_policy") in ("reject", "quarantine"))
        spf_ok   = sum(1 for r in resultaten if r.get("spf_record"))
        sec_txt  = sum(1 for r in resultaten if r.get("security_txt") == "found")

        log.info("\n📊 Samenvatting:")
        log.info(f"   Microsoft 365:       {ms_count:>4}")
        log.info(f"   Google Workspace:    {g_count:>4}")
        log.info(f"   SPF aanwezig:        {spf_ok:>4}")
        log.info(f"   DMARC (q/r):         {dmarc_ok:>4}")
        log.info(f"   security.txt:        {sec_txt:>4}")


if __name__ == "__main__":
    main()

